"""
文档解析工具

支持解析 PDF、Word、Excel 文档，绑定 MinIO 文件路径白名单。
"""
import io
import os
import logging
from typing import Dict, Any, Optional
from pathlib import Path

from app.agent.tools.base import BaseTool, ToolResult
from app.config import get_settings
from app.utils.path_utils import get_data_dir

logger = logging.getLogger(__name__)
settings = get_settings()


class DocumentParserTool(BaseTool):
    """
    文档解析工具
    
    功能：
    - 解析 PDF 文档
    - 解析 Word 文档（.docx）
    - 解析 Excel 表格（.xlsx）
    
    安全机制：
    - 文件路径白名单：只允许从 MinIO 或指定目录读取
    - 禁止任意路径读取
    - 文件大小限制
    """
    
    # 允许的文件类型
    ALLOWED_EXTENSIONS = {
        '.pdf': 'PDF 文档',
        '.docx': 'Word 文档',
        '.xlsx': 'Excel 表格',
        '.txt': '文本文件',
    }
    
    # 最大文件大小（50MB）
    MAX_FILE_SIZE = 50 * 1024 * 1024
    
    # 允许的目录白名单
    ALLOWED_DIRECTORIES = []
    
    def __init__(self):
        """初始化文档解析工具"""
        super().__init__(
            name="document_parser",
            description="解析 PDF、Word、Excel 等文档，提取文本内容"
        )
        self.timeout = 60  # 文档解析可能较慢
        
        # 初始化允许的目录
        self._init_allowed_directories()
    
    def _init_allowed_directories(self):
        """初始化允许的目录白名单"""
        # 数据目录
        data_dir = get_data_dir()
        self.ALLOWED_DIRECTORIES.append(str(data_dir))
        
        # MinIO 数据目录（如果配置了）
        minio_path = os.getenv('MINIO_DATA_PATH', '/data/minio')
        if os.path.exists(minio_path):
            self.ALLOWED_DIRECTORIES.append(minio_path)
    
    def validate_params(self, **kwargs) -> bool:
        """
        验证参数
        
        Args:
            file_path: 文件路径
            file_type: 文件类型（可选）
        
        Returns:
            bool: 参数是否有效
        
        Raises:
            ValueError: 参数无效
        """
        file_path = kwargs.get('file_path')
        
        if not file_path:
            raise ValueError("Parameter 'file_path' is required")
        
        if not isinstance(file_path, str):
            raise ValueError("Parameter 'file_path' must be a string")
        
        # 检查文件扩展名
        ext = Path(file_path).suffix.lower()
        if ext not in self.ALLOWED_EXTENSIONS:
            raise ValueError(
                f"Unsupported file type: {ext}. "
                f"Allowed: {list(self.ALLOWED_EXTENSIONS.keys())}"
            )
        
        # 检查路径是否在白名单中
        if not self._is_path_allowed(file_path):
            raise ValueError(
                f"File path not in allowed directories. "
                f"Allowed: {self.ALLOWED_DIRECTORIES}"
            )
        
        return True
    
    def _is_path_allowed(self, file_path: str) -> bool:
        """
        检查路径是否在白名单中
        
        Args:
            file_path: 文件路径
        
        Returns:
            bool: 是否允许
        """
        # 解析绝对路径
        abs_path = os.path.abspath(file_path)
        
        # 检查是否在允许的目录中
        for allowed_dir in self.ALLOWED_DIRECTORIES:
            if abs_path.startswith(allowed_dir):
                return True
        
        return False
    
    async def execute(self, **kwargs) -> ToolResult:
        """
        执行文档解析
        
        Args:
            file_path: 文件路径
            file_type: 文件类型（可选，自动检测）
        
        Returns:
            ToolResult: 解析结果
        """
        file_path = kwargs.get('file_path')
        
        try:
            # 检查文件是否存在
            if not os.path.exists(file_path):
                return ToolResult(
                    success=False,
                    output=None,
                    error=f"File not found: {file_path}"
                )
            
            # 检查文件大小
            file_size = os.path.getsize(file_path)
            if file_size > self.MAX_FILE_SIZE:
                return ToolResult(
                    success=False,
                    output=None,
                    error=f"File too large: {file_size} bytes (max {self.MAX_FILE_SIZE} bytes)"
                )
            
            # 获取文件扩展名
            ext = Path(file_path).suffix.lower()
            
            # 根据文件类型选择解析方法
            if ext == '.pdf':
                content = await self._parse_pdf(file_path)
            elif ext == '.docx':
                content = await self._parse_word(file_path)
            elif ext == '.xlsx':
                content = await self._parse_excel(file_path)
            elif ext == '.txt':
                content = await self._parse_text(file_path)
            else:
                return ToolResult(
                    success=False,
                    output=None,
                    error=f"Unsupported file type: {ext}"
                )
            
            return ToolResult(
                success=True,
                output=content,
                metadata={
                    "file_path": file_path,
                    "file_type": ext,
                    "file_size": file_size,
                    "content_length": len(content)
                }
            )
            
        except Exception as e:
            logger.error(f"Document parser tool error: {e}")
            return ToolResult(
                success=False,
                output=None,
                error=f"Document parsing failed: {str(e)}"
            )
    
    async def _parse_pdf(self, file_path: str) -> str:
        """
        解析 PDF 文件
        
        Args:
            file_path: PDF 文件路径
        
        Returns:
            提取的文本内容
        """
        try:
            # 尝试使用 PyPDF2
            from PyPDF2 import PdfReader
            
            reader = PdfReader(file_path)
            text_parts = []
            
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)
            
            return "\n\n".join(text_parts)
            
        except ImportError:
            logger.warning("PyPDF2 not installed, using fallback method")
            # 如果没有安装 PyPDF2，返回提示信息
            return f"[PDF 文档解析需要安装 PyPDF2 库]\n文件路径: {file_path}"
    
    async def _parse_word(self, file_path: str) -> str:
        """
        解析 Word 文件
        
        Args:
            file_path: Word 文件路径
        
        Returns:
            提取的文本内容
        """
        try:
            # 尝试使用 python-docx
            from docx import Document
            
            doc = Document(file_path)
            text_parts = []
            
            for paragraph in doc.paragraphs:
                if paragraph.text:
                    text_parts.append(paragraph.text)
            
            return "\n\n".join(text_parts)
            
        except ImportError:
            logger.warning("python-docx not installed, using fallback method")
            return f"[Word 文档解析需要安装 python-docx 库]\n文件路径: {file_path}"
    
    async def _parse_excel(self, file_path: str) -> str:
        """
        解析 Excel 文件
        
        Args:
            file_path: Excel 文件路径
        
        Returns:
            提取的文本内容
        """
        try:
            # 尝试使用 openpyxl
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path, read_only=True, data_only=True)
            text_parts = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text_parts.append(f"=== Sheet: {sheet_name} ===")
                
                for row in sheet.iter_rows(values_only=True):
                    # 过滤空值
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    if any(row_data):  # 至少有一个非空值
                        text_parts.append(" | ".join(row_data))
            
            return "\n".join(text_parts)
            
        except ImportError:
            logger.warning("openpyxl not installed, using fallback method")
            return f"[Excel 表格解析需要安装 openpyxl 库]\n文件路径: {file_path}"
    
    async def _parse_text(self, file_path: str) -> str:
        """
        解析文本文件
        
        Args:
            file_path: 文本文件路径
        
        Returns:
            文本内容
        """
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    
    def _get_parameters_schema(self) -> Dict[str, Any]:
        """获取参数 schema"""
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "文档文件路径（必须在允许的目录中）"
                },
                "file_type": {
                    "type": "string",
                    "description": "文件类型（可选，自动检测）",
                    "enum": ["pdf", "docx", "xlsx", "txt"]
                }
            },
            "required": ["file_path"]
        }